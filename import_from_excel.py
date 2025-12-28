
from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from main import Base, Tank, Parameter, TankParameter, Sample, SampleValue, DEFAULT_UNITS

EXCEL_PATH = os.environ.get("REEF_EXCEL_PATH", os.path.join(os.path.dirname(__file__), "..", "..", "Tank Parameters Sheet.xlsx"))
DB_PATH = os.environ.get("REEF_DB_PATH", os.path.join(os.path.dirname(__file__), "reef.db"))

TANK_SHEETS = ["Display Tank", "Frag Tank", "Coral Bays", "4m Tank", "Cade", "Red Sea"]

PARAM_DECIMALS = {
    "Alkalinity": 2,
    "Calcium": 0,
    "Magnesium": 0,
    "Nitrate": 2,
    "Phosphate": 3,
    "Potassium": 0,
    "Salinity": 1,
    "Temperature": 1,
}

def parse_tank_sheet(ws):
    # Find parameter headers
    headers = []
    for c in range(1, 40):
        v = ws.cell(1, c).value
        if v is None:
            continue
        headers.append((c, v))

    param_cols = []
    for c, v in headers:
        if v == "Tank Volume":
            break
        if isinstance(v, (int, float)):
            continue
        param_cols.append((c, v))

    # volume
    volume = None
    for c, v in headers:
        if v == "Tank Volume":
            volume = ws.cell(1, c + 1).value
            break

    # rows
    rows = []
    blanks = 0
    for r in range(2, ws.max_row + 1):
        ts = ws.cell(r, param_cols[0][0]).value
        if ts is None:
            blanks += 1
            if blanks >= 8:
                break
            continue
        blanks = 0
        rec = {"taken_at": ts}
        for c, name in param_cols[1:]:
            rec[name] = ws.cell(r, c).value
        rows.append(rec)

    return [name for _, name in param_cols], volume, rows

def main():
    if not os.path.exists(EXCEL_PATH):
        raise SystemExit(f"Excel not found at {EXCEL_PATH}")

    engine = create_engine(f"sqlite:///{DB_PATH}", connect_args={"check_same_thread": False})
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(engine)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    with SessionLocal() as db:
        # parameters
        param_by_name: Dict[str, Parameter] = {}
        for sheet in TANK_SHEETS:
            ws = wb[sheet]
            cols, vol, rows = parse_tank_sheet(ws)
            for p in cols[1:]:
                if p in param_by_name:
                    continue
                obj = db.execute(select(Parameter).where(Parameter.name == p)).scalar_one_or_none()
                if not obj:
                    obj = Parameter(
                        name=p,
                        unit=DEFAULT_UNITS.get(p),
                        decimals=PARAM_DECIMALS.get(p),
                    )
                    db.add(obj)
                    db.flush()
                param_by_name[p] = obj

        # tanks + tank parameters
        tank_by_name: Dict[str, Tank] = {}
        for sheet in TANK_SHEETS:
            ws = wb[sheet]
            cols, vol, rows = parse_tank_sheet(ws)
            tank = db.execute(select(Tank).where(Tank.name == sheet)).scalar_one_or_none()
            if not tank:
                tank = Tank(name=sheet, volume_l=float(vol) if vol is not None else None)
                db.add(tank)
                db.flush()
            else:
                tank.volume_l = float(vol) if vol is not None else tank.volume_l
            tank_by_name[sheet] = tank

            # link parameters in displayed order
            for i, p_name in enumerate(cols[1:], start=1):
                p = param_by_name[p_name]
                link = db.execute(
                    select(TankParameter).where(TankParameter.tank_id == tank.id, TankParameter.parameter_id == p.id)
                ).scalar_one_or_none()
                if not link:
                    db.add(TankParameter(tank_id=tank.id, parameter_id=p.id, sort_order=i))

        db.commit()

        # samples
        for sheet in TANK_SHEETS:
            ws = wb[sheet]
            cols, vol, rows = parse_tank_sheet(ws)
            tank = tank_by_name[sheet]

            for rec in rows:
                taken_at = rec["taken_at"]
                # normalize excel datetimes
                if isinstance(taken_at, datetime):
                    dt_local = taken_at
                else:
                    # try parsing string
                    try:
                        dt_local = datetime.fromisoformat(str(taken_at))
                    except Exception:
                        continue

                sample = Sample(tank_id=tank.id, taken_at=dt_local, notes=None)
                db.add(sample)
                db.flush()

                any_val = False
                for p_name in cols[1:]:
                    raw = rec.get(p_name)
                    if raw is None or raw == "":
                        continue
                    try:
                        val = float(raw)
                    except Exception:
                        continue
                    db.add(SampleValue(sample_id=sample.id, parameter_id=param_by_name[p_name].id, value=val))
                    any_val = True

                if not any_val:
                    db.rollback()
                else:
                    db.commit()

    print("✅ Import complete.")
    print(f"DB: {DB_PATH}")

if __name__ == "__main__":
    main()
